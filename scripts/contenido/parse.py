# -*- coding: utf-8 -*-
"""Extrae preguntas de los Excel de SIMA CHECK.

Convenciones del Excel (verificadas sobre los 5 archivos):
  - El enunciado va en la celda C3 de su fila, o en el texto de un cuadro de
    dibujo, y siempre arranca con "N - ".
  - Opciones de texto: grilla 2x2 en las columnas 4 y 7 (dos filas), o una
    sola columna 5 para Verdadero/Falso y A/B.
  - La correcta esta pintada de verde (FF00B050).
  - Opciones con imagen: los dibujos anclados en columnas 3..9; la correcta es
    la que tiene la CELDA VACIA verde en su misma (fila, columna).
"""
import json, re, sys, zipfile
import openpyxl
from xml.etree import ElementTree as ET

NS = {
    'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
    'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
    'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
}
VERDE = '00B050'
RE_NUM = re.compile(r'^\s*(\d+)\s*[-–.]\s*(.*)$', re.S)


def es_verde(cell):
    if not cell.fill.patternType:
        return False
    fg = cell.fill.fgColor
    return fg.type == 'rgb' and str(fg.rgb).endswith(VERDE)


def leer_dibujos(path):
    """[(fila, col, [lineas de texto], [(offset_x, nombre_imagen)])] ordenado."""
    z = zipfile.ZipFile(path)
    salida = []
    for nombre in z.namelist():
        if not (nombre.startswith('xl/drawings/drawing') and nombre.endswith('.xml')):
            continue
        rels_path = nombre.replace('drawings/', 'drawings/_rels/') + '.rels'
        try:
            rels = {e.get('Id'): e.get('Target').split('/')[-1]
                    for e in ET.fromstring(z.read(rels_path))}
        except KeyError:
            rels = {}
        for anchor in ET.fromstring(z.read(nombre)):
            desde = anchor.find('xdr:from', NS)
            if desde is None:
                continue
            fila = int(desde.find('xdr:row', NS).text) + 1
            col = int(desde.find('xdr:col', NS).text) + 1
            textos = [''.join(t.text or '' for t in p.findall('.//a:t', NS))
                      for p in anchor.findall('.//a:p', NS)]
            textos = [t.strip() for t in textos if t.strip()]
            # Las imagenes se ordenan por su x dentro del grupo: es lo unico que
            # distingue la opcion "A" de la "B" cuando las dos cuelgan del mismo
            # anchor (preguntas 42-44 de Basico).
            pics = []
            for pic in anchor.findall('.//xdr:pic', NS):
                blip = pic.find('.//a:blip', NS)
                off = pic.find('.//a:xfrm/a:off', NS)
                x = int(off.get('x')) if off is not None else 0
                pics.append((x, rels.get(blip.get('{%s}embed' % NS['r']), '')))
            pics.sort()
            if textos or pics:
                salida.append((fila, col, textos, pics))
    salida.sort(key=lambda e: (e[0], e[1]))
    return salida


def parsear(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.worksheets[0]
    dibujos = leer_dibujos(path)

    # --- Eventos por fila -----------------------------------------------
    # enunciado / opcion-texto / imagen / verde-vacio, todos ordenados por fila.
    eventos = []
    for fila in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=fila, column=col)
            valor = str(cell.value).strip() if cell.value is not None else ''
            if valor:
                m = RE_NUM.match(valor)
                if col == 3 and m:
                    eventos.append((fila, col, 'enunciado', (int(m.group(1)), m.group(2).strip())))
                else:
                    eventos.append((fila, col, 'opcion', (valor, es_verde(cell))))
            elif es_verde(cell):
                eventos.append((fila, col, 'verde-vacio', None))
    for fila, col, textos, pics in dibujos:
        # El enunciado no siempre es la primera linea del cuadro: en las
        # preguntas "A o B" de Basico las etiquetas A/B vienen antes.
        idx = next((i for i, t in enumerate(textos) if RE_NUM.match(t)), None)
        if idx is not None:
            m = RE_NUM.match(textos[idx])
            cuerpo = '\n'.join([m.group(2).strip()] + textos[idx + 1:]).strip()
            eventos.append((fila, col, 'enunciado', (int(m.group(1)), cuerpo)))
            resto = textos[:idx]
        else:
            resto = textos
        for i, (_, img) in enumerate(pics):
            eventos.append((fila, col + i * 0.001, 'imagen', (img, resto)))
    eventos.sort(key=lambda e: (e[0], e[1]))

    # --- Agrupar en preguntas -------------------------------------------
    preguntas, actual = [], None
    for fila, col, tipo, datos in eventos:
        if tipo == 'enunciado':
            actual = {'n': datos[0], 'texto': datos[1], 'opciones': [],
                      'imagenes': [], 'verdes': []}
            preguntas.append(actual)
            continue
        if actual is None:
            continue  # encabezados y logos previos a la pregunta 1
        if tipo == 'opcion':
            actual['opciones'].append({'fila': fila, 'col': col,
                                       'texto': datos[0], 'verde': datos[1]})
        elif tipo == 'imagen':
            actual['imagenes'].append({'fila': fila, 'col': int(col), 'img': datos[0]})
        else:
            actual['verdes'].append((fila, int(col)))
    return preguntas


if __name__ == '__main__':
    salida = {}
    for path in sys.argv[1:]:
        salida[path] = parsear(path)
    print(json.dumps(salida, ensure_ascii=False, indent=1))
