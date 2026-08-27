# -*- coding: utf-8 -*-
"""Genera prisma/seed-data/catalogos-nomina.ts desde el Excel de nomina.

Solo se lleva los CATALOGOS (nombres de puestos y de centros de costo). Los 264
empleados con nombre y DNI NO se tocan: son PII y los carga el admin desde el
import de Excel del backoffice.
"""
import os
import openpyxl

RAIZ = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
ORIGEN = os.path.join(RAIZ, 'docs', 'Puestos y listado de personal.xlsx')
DESTINO = os.path.join(RAIZ, 'sima-training-api', 'prisma', 'seed-data')

CABECERA = """// ARCHIVO GENERADO — no editar a mano.
//
// Catálogos de nómina de Ingeniería SIMA, sacados de
// `docs/Puestos y listado de personal.xlsx`: la hoja "Listado de Puestos" y los
// valores distintos de la columna "Dependencia" de la hoja de nómina (que es
// como la empresa llama a lo que acá es centro de costo).
//
// Del Excel de nómina se lleva SOLO esto. Las 264 filas de personas —legajo,
// DNI y apellido y nombre— son PII y no entran a ningún archivo versionado:
// se cargan desde el import de Excel del backoffice.

"""


def literal(valor):
    return "'" + valor.replace('\\', '\\\\').replace("'", "\\'") + "'"


def bloque(nombre, valores, comentario):
    lineas = [comentario, 'export const ' + nombre + ': string[] = [']
    for valor in valores:
        lineas.append('  ' + literal(valor) + ',')
    lineas.append('];')
    return '\n'.join(lineas)


def main():
    wb = openpyxl.load_workbook(ORIGEN, data_only=True)

    hoja_puestos = wb['Listado de Puestos']
    puestos = []
    for fila in range(3, hoja_puestos.max_row + 1):
        valor = hoja_puestos.cell(row=fila, column=2).value
        if valor and str(valor).strip():
            puestos.append(' '.join(str(valor).split()))

    hoja_nomina = wb['Nómina de personal ']
    centros = []
    for fila in range(2, hoja_nomina.max_row + 1):
        valor = hoja_nomina.cell(row=fila, column=4).value
        if valor and str(valor).strip():
            nombre = ' '.join(str(valor).split())
            if nombre not in centros:
                centros.append(nombre)
    centros.sort()

    texto = CABECERA
    texto += bloque('PUESTOS', puestos,
                    '// Los ' + str(len(puestos)) + ' puestos de la hoja "Listado de Puestos", en su orden.')
    texto += '\n\n'
    texto += bloque('CENTROS_COSTO', centros,
                    '// Los ' + str(len(centros)) + ' valores distintos de "Dependencia" en la nómina.')
    texto += '\n'

    os.makedirs(DESTINO, exist_ok=True)
    with open(os.path.join(DESTINO, 'catalogos-nomina.ts'), 'w', encoding='utf8', newline='\n') as fh:
        fh.write(texto)
    print(str(len(puestos)) + ' puestos, ' + str(len(centros)) + ' centros de costo')


if __name__ == '__main__':
    main()
